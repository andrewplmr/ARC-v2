import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from engine.utils import load_config


cfg = load_config()

# ============================================================
# Constants
# ============================================================

# Columns that should never be exposed to the client
INTERNAL_COLUMNS = {
    "ref_norm",
    "match_key",
    "amount_cent",
    "fuzzy_status",
    "polarity",
    "currency",
    "amount_gbp",

    # Internal reasoning fields
    "match_type",
    "resolution_status",
    "reason_code",
    "variance_amount",
}


# User-facing column renames
COLUMN_RENAMES = {
    "final_status": "Status",
    "source": "Source",
    "date": "Date",
    "amount": "Amount (£)",
    "reference": "Reference",
    "Match Reason": "Match Reason",
}


# Client-friendly status translations
CLIENT_STATUS_MAP = {
    "Matched": "Matched",
    "FuzzyMatched": "Partially Matched",
    "PartialMatched": "Partially Matched",
    "Unmatched": "Unmatched",
}

# Cell fills by status
STATUS_FILLS = {
    "Matched": PatternFill("solid", start_color="C6EFCE"),
    "PartialMatched": PatternFill("solid", start_color="FFF2CC"),
    "Unmatched": PatternFill("solid", start_color="F4CCCC"),
}

# Reusable thin border
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Fixed-width margin columns
MARGIN_COLUMNS = {
    "A": 2.8,
    "D": 2.8,
}

# ============================================================
# Helper Utilities
# ============================================================

def _auto_width_range(
    ws: Worksheet,
    min_col: int,
    max_col: int,
    min_row: int,
    max_row: int,
):
    """
    Auto-sizes columns using ONLY values within a defined cell range.
    Explicitly ignores headers, watermarks, and margin columns.
    """
    for col_idx in range(min_col, max_col + 1):
        max_len = 0

        for row in range(min_row, max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        if max_len > 0:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 3,
                40,
            )


def _enforce_min_width(ws: Worksheet, columns, min_width: int = 18):
    """
    Ensures columns respect a minimum width without overriding auto-sizing.
    """
    for col_idx in columns:
        letter = get_column_letter(col_idx)
        current = ws.column_dimensions[letter].width

        if current is None or current < min_width:
            ws.column_dimensions[letter].width = min_width


def _draw_header_separator(ws: Worksheet, start_col="B", end_col="I", row=1):
    """
    Draws a horizontal divider line under the main dashboard header.
    """
    thin = Side(style="thin", color="000000")

    for col in range(
        ws[f"{start_col}{row}"].column,
        ws[f"{end_col}{row}"].column + 1,
    ):
        ws.cell(row=row, column=col).border = Border(bottom=thin)


def _hide_gridlines(ws: Worksheet):
    """Removes default Excel gridlines for a clean dashboard look."""
    ws.sheet_view.showGridLines = False


def _prettify(df: pd.DataFrame) -> pd.DataFrame:
    """
    Applies all client-facing formatting rules to a dataframe:
    - Column renames
    - Title-casing
    - Status translation
    - Source capitalization
    """
    df = df.copy()

    # Rename + title-case columns
    df.rename(columns=COLUMN_RENAMES, inplace=True)
    df.columns = [c.replace("_", " ").title() for c in df.columns]

    # Translate internal statuses
    if "Status" in df.columns:
        df["Status"] = df["Status"].map(lambda x: CLIENT_STATUS_MAP.get(x, x))

    # Capitalise source names
    if "Source" in df.columns:
        df["Source"] = df["Source"].astype(str).str.title()

    return df


def _split_internal(df: pd.DataFrame):
    """
    Separates client-visible columns from internal-only metadata.
    """
    visible, internal = [], []

    for col in df.columns:
        normalised = col.lower().replace(" ", "_")
        (internal if normalised in INTERNAL_COLUMNS else visible).append(col)

    return df[visible], df[internal]


def _apply_zebra_rows(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
):
    """Applies alternating row shading for readability."""
    fill = PatternFill("solid", start_color="EAEAEA")

    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            for c in range(start_col, end_col + 1):
                ws.cell(r, c).fill = fill


def _color_status_column(
    ws: Worksheet,
    col: int,
    start_row: int,
    end_row: int,
):
    """Applies colour coding and centering to the Status column."""
    for r in range(start_row, end_row + 1):
        cell = ws.cell(r, col)
        value = str(cell.value).strip()

        if value == "Matched":
            cell.fill = STATUS_FILLS["Matched"]
        elif value == "Partially Matched":
            cell.fill = STATUS_FILLS["PartialMatched"]
        elif value == "Unmatched":
            cell.fill = STATUS_FILLS["Unmatched"]
        else:
            continue

        cell.alignment = Alignment(horizontal="center")


def _sort_transactions(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sorts transactions by date descending when available.
    """
    return df.sort_values(by="date", ascending=False) if "date" in df.columns else df

# ============================================================
# Transaction Table Writer
# ============================================================

def _write_transaction_table(
    ws: Worksheet,
    df: pd.DataFrame,
    start_row: int,
    start_col: int = 5,
    write_header: bool = True,
):
    """
    Writes a transaction table to the worksheet.

    - Header is written only once
    - Always leaves a single blank row after the table
    - Returns the next safe starting row
    """
    df = _prettify(df)
    visible, _ = _split_internal(df)

    header_row = start_row if write_header else None
    data_start = start_row + 1 if write_header else start_row
    data_end = data_start + len(visible) - 1
    end_col = start_col + len(visible.columns) - 1

    # Locate Date column for alignment
    date_col_idx = (
        start_col + visible.columns.get_loc("Date")
        if "Date" in visible.columns
        else None
    )

    # Header
    if write_header:
        for i, col_name in enumerate(visible.columns):
            cell = ws.cell(header_row, start_col + i, col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="000000")
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Data rows
    for r_idx, row in enumerate(visible.itertuples(index=False), start=data_start):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = THIN_BORDER

            if date_col_idx and c_idx == date_col_idx:
                cell.alignment = Alignment(horizontal="center")

    _apply_zebra_rows(ws, data_start, data_end, start_col, end_col)

    if "Status" in visible.columns:
        status_col = start_col + visible.columns.get_loc("Status")
        _color_status_column(ws, status_col, data_start, data_end)

    return data_end + 2

# ============================================================
# Dashboard Sheet Builder
# ============================================================

def write_dashboard_sheet(
    wb: Workbook,
    summary_df: pd.DataFrame,
    matched_df: pd.DataFrame,
    partial_df: pd.DataFrame,
    unmatched_df: pd.DataFrame,
    client_name: str,
):
    ws = wb.create_sheet("Dashboard")
    _hide_gridlines(ws)

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------
    ws.row_dimensions[1].height = 50
    ws["B1"] = client_name
    ws["B1"].font = Font(size=22, bold=True)
    ws["B1"].alignment = Alignment(vertical="center")

    ws["I1"] = "Automated Reconciliation Core"
    ws["I1"].alignment = Alignment(horizontal="right")

    _draw_header_separator(ws)

    # --------------------------------------------------------
    # Summary Metrics
    # --------------------------------------------------------
    total = len(summary_df)
    matched = (summary_df["final_status"] == "Matched").sum()
    partial = summary_df["final_status"].isin(["FuzzyMatched"]).sum()
    unmatched = (summary_df["final_status"] == "Unmatched").sum()
    rate = matched / total if total else 0

    rows = [
        ("Total Transactions", total),
        ("Matched", matched),
        ("Partially Matched", partial),
        ("Unmatched", unmatched),
        ("Match Rate (%)", rate),
        ("Total Amount (£)", summary_df["amount"].sum()),
        (
            "Unmatched Amount (£)",
            summary_df.loc[summary_df["final_status"] == "Unmatched", "amount"].sum(),
        ),
    ]

    ws["B3"], ws["C3"] = "Metric", "Value"
    for cell in ("B3", "C3"):
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill("solid", start_color="000000")

    for i, (label, value) in enumerate(rows, start=4):
        ws[f"B{i}"] = label
        ws[f"C{i}"] = value
        ws[f"C{i}"].alignment = Alignment(horizontal="left")

    ws["C8"].number_format = "0.00%"
    ws["C9"].number_format = "£#,##0.00"
    ws["C10"].number_format = "£#,##0.00"

    for r in range(3, 3 + len(rows) + 1):
        for c in (2, 3):
            ws.cell(r, c).border = THIN_BORDER

    _apply_zebra_rows(ws, 4, 3 + len(rows), 2, 3)

    # --------------------------------------------------------
    # Transaction Tables
    # --------------------------------------------------------
    current_row = 3

    matched_df = _sort_transactions(matched_df)
    partial_df = _sort_transactions(partial_df)
    unmatched_df = _sort_transactions(unmatched_df)

    current_row = _write_transaction_table(ws, matched_df, current_row, write_header=True)
    current_row = _write_transaction_table(ws, partial_df, current_row, write_header=False)
    current_row = _write_transaction_table(ws, unmatched_df, current_row, write_header=False)

    final_table_row = current_row - 2

    # --------------------------------------------------------
    # Footer / Watermark
    # --------------------------------------------------------
    today = datetime.today().strftime("%Y-%m-%d")
    ws["B18"] = "Prepared automatically by ARC Solutions"
    ws["B19"] = f"Generated on: {today}"

    # --------------------------------------------------------
    # Column Sizing
    # --------------------------------------------------------
    _auto_width_range(ws, 2, 3, 3, 3 + len(rows))
    _auto_width_range(ws, 5, ws.max_column, 3, final_table_row)

    _enforce_min_width(ws, [2, 3], min_width=18)
    _enforce_min_width(ws, range(5, ws.max_column + 1), min_width=18)

    for col, width in MARGIN_COLUMNS.items():
        ws.column_dimensions[col].width = width

# ============================================================
# Workbook Writer
# ============================================================

def write_styled_workbook(
    summary_df: pd.DataFrame,
    reconciled_df: pd.DataFrame,
    partial_df: pd.DataFrame,
    unmatched_df: pd.DataFrame,
    client_name: str,
    out_path: str,
):
    wb = Workbook()
    wb.remove(wb.active)

    write_dashboard_sheet(
        wb,
        summary_df,
        reconciled_df,
        partial_df,
        unmatched_df,
        client_name,
    )

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb.save(out_path)
    logging.info(f"Excel written to {out_path}")

    return out_path
